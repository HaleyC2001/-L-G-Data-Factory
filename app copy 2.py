import streamlit as st
import pandas as pd
import numpy as np
import io
from datetime import date
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


# adding comments to test git branch functionality

def clean_sheet_name(name, fallback="Sheet"):
    name = str(name).strip()
    for ch in [":", "/", "\\", "?", "*", "[", "]"]:
        name = name.replace(ch, "")
    name = name.strip("'")
    if not name:
        name = fallback
    return name[:31]


st.set_page_config(page_title="Weekly Update Report Tool", page_icon="📊", layout="centered")

st.title("📊 Weekly Update Report Tool")
st.markdown("Upload your three Excel files, set your target numbers, and download the report — no coding needed.")
st.divider()

# ── Step 1: File Uploads ──────────────────────────────────────────────────────
st.subheader("Step 1 · Upload Your Excel Files")
col1, col2, col3 = st.columns(3)
with col1:
    students_file = st.file_uploader("🎒 Students File", type=["xlsx", "xls"], key="students")
with col2:
    adults_file = st.file_uploader("👪 Adults File", type=["xlsx", "xls"], key="adults")
with col3:
    all_file = st.file_uploader("📋 Grant Level File", type=["xlsx", "xls"], key="all")

st.divider()

# ── Step 1.5: Institution Name ────────────────────────────────────────────────
institution_name = st.text_input("🏫 Institution Name (used in output filename)", value="Institution")

st.divider()

# ── Step 2: Target values ─────────────────────────────────────────────────────
st.subheader("Step 2 · Set Target Enrollment per Site")
st.info("Enter the target number of students for each site — in the same order as they appear in your Students file. Don't include the Total row.")

if "num_sites" not in st.session_state:
    st.session_state.num_sites = 3

col_add, col_remove = st.columns([1, 1])
with col_add:
    if st.button("➕ Add a site"):
        st.session_state.num_sites += 1
with col_remove:
    if st.button("➖ Remove last site") and st.session_state.num_sites > 1:
        st.session_state.num_sites -= 1

target_values = []
defaults = [152, 200, 100]
for i in range(st.session_state.num_sites):
    default = defaults[i] if i < len(defaults) else 100
    val = st.number_input(f"Site {i+1} — Target # of students", min_value=1, value=default, key=f"target_{i}")
    target_values.append(val)

st.divider()

# ── Step 3: Generate ──────────────────────────────────────────────────────────
st.subheader("Step 3 · Generate Report")

all_uploaded = students_file and adults_file and all_file
if not all_uploaded:
    st.warning("Please upload all three files above before generating.")

if st.button("🚀 Generate Report", disabled=not all_uploaded, type="primary", use_container_width=True):
    with st.status("Running pipeline…", expanded=True) as status:

        try:
            st.write("📂 Reading uploaded files…")
            students_bytes = students_file.read()
            students_sheets = pd.read_excel(io.BytesIO(students_bytes), sheet_name=None)
            adults_bytes = adults_file.read()
            all_bytes = all_file.read()

            # ── Student Summary Statistics ────────────────────────────────────
            st.write("🍀 Processing Student Summary Statistics…")

            # Find the Participants By Hour Band sheet (sheet name may be truncated)
            part_by_hour_key = None
            for sheet_name in students_sheets.keys():
                if 'Participants By Hour Band' in sheet_name:
                    part_by_hour_key = sheet_name
                    break
            if part_by_hour_key is None:
                raise ValueError("Could not find 'Participants By Hour Band' sheet in Students file")

            df_part_by_hour = students_sheets[part_by_hour_key].copy()

            # Use the 2nd row (index 1) as column names
            df_part_by_hour.columns = df_part_by_hour.iloc[1]
            # Drop the first two rows and reset index
            df_part_by_hour = df_part_by_hour.iloc[2:].reset_index(drop=True)

            # Filter rows: keep only rows where Institution is empty AND Site is non-empty
            rows_to_drop = []
            for idx, row in df_part_by_hour.iterrows():
                institution = row['Institution']
                site = row['Site']

                is_empty_institution = pd.isna(institution) or str(institution).strip() == "NaN"
                is_empty_site = pd.isna(site) or str(site).strip() == "NaN"

                if (is_empty_institution and is_empty_site) or not is_empty_institution:
                    rows_to_drop.append(idx)

            df_part_by_hour = df_part_by_hour.drop(index=rows_to_drop).reset_index(drop=True)

            # Daily Site Attendance Summary
            daily_site_att = students_sheets['Daily Site Attendance Summary'].copy()
            daily_site_att.columns = daily_site_att.iloc[2]
            daily_site_att = daily_site_att.iloc[3:]
            daily_site_att.columns.name = None
            daily_site_att = daily_site_att.reset_index(drop=True)
            daily_site_att = daily_site_att[['Total']].iloc[:-1]
            daily_site_att['Total'] = daily_site_att['Total'].str.extract(r'(\d+\.?\d*)')

            all_cols = ['0 Hours', 'Less Than 15 Hours', '15-44 Hours', '45-89 Hours', '90-179 Hours', '180-269 Hours', '270+ Hours']
            served_cols = ['Less Than 15 Hours', '15-44 Hours', '45-89 Hours', '90-179 Hours', '180-269 Hours', '270+ Hours']
            plus15_cols = ['15-44 Hours', '45-89 Hours', '90-179 Hours', '180-269 Hours', '270+ Hours']
            plus90_cols = ['90-179 Hours', '180-269 Hours', '270+ Hours']
            existing_cols = df_part_by_hour.columns.tolist()
            all_cols = [c for c in all_cols if c in existing_cols]
            served_cols = [c for c in served_cols if c in existing_cols]
            plus15_cols = [c for c in plus15_cols if c in existing_cols]
            plus90_cols = [c for c in plus90_cols if c in existing_cols]

            df_part_by_hour[all_cols] = df_part_by_hour[all_cols].apply(pd.to_numeric, errors='coerce').fillna(0)

            df_totals = pd.DataFrame({
                '[Total # Enrolled]': df_part_by_hour[all_cols].sum(axis=1),
                '[Total # Served]': df_part_by_hour[served_cols].sum(axis=1),
                '[Total # 15+]': df_part_by_hour[plus15_cols].sum(axis=1),
                '[Total # 90+]': df_part_by_hour[plus90_cols].sum(axis=1),
            })
            df_totals.insert(0, '[Target # of students served]', target_values)

            daily_site_att['Total'] = daily_site_att['Total'].astype(int)
            daily_site_att = daily_site_att.rename(columns={'Total': 'Avg. # of Students Per Day'})
            df_totals.insert(3, 'Avg. # of Students Per Day', daily_site_att['Avg. # of Students Per Day'].values)

            school_names = [
                row['Site'] for _, row in df_part_by_hour.iterrows()
                if pd.notna(row['Site']) and row['Site'] != 'Subtotal' and row['Site'] != 'Total'
            ]
            df_totals.insert(0, 'School', school_names)

            total_row = pd.DataFrame(df_totals.iloc[:, 1:].sum()).T
            total_row.insert(0, 'School', 'Total')
            df_totals = pd.concat([df_totals, total_row], ignore_index=True)

            df_totals['# of students 15+ hrs total (% of Target)'] = (
                df_totals['[Total # 15+]'].astype(int).astype(str) + " (" +
                ((df_totals['[Total # 15+]'] / df_totals['[Target # of students served]']) * 100)
                .round().astype(int).astype(str) + "%)"
            )
            df_totals['# of students 90+ hrs total (% of Target)'] = (
                df_totals['[Total # 90+]'].astype(int).astype(str) + " (" +
                ((df_totals['[Total # 90+]'] / df_totals['[Target # of students served]']) * 100)
                .round().astype(int).astype(str) + "%)"
            )
            df_totals = df_totals.drop(columns=['[Total # 15+]', '[Total # 90+]']).reset_index(drop=True)

            # ── Family Component ──────────────────────────────────────────────
            st.write("🌷 Processing Family Component…")
            df_hours = pd.read_excel(io.BytesIO(adults_bytes), sheet_name="Participant Attendance Hours", skiprows=2)
            df_hours['HoursPresent'] = pd.to_numeric(df_hours['HoursPresent'], errors='coerce')
            df_hours['ParticipantId'] = df_hours['ParticipantId'].astype(str).str.replace(r'\.0$', '', regex=True)
            df_active = df_hours[(df_hours['HoursPresent'] > 0) & (df_hours['ParticipantId'].str.len() != 9)]
            result = df_active.groupby('Site')['ParticipantId'].nunique().reset_index()
            result.rename(columns={'ParticipantId': 'Parents Served (Total)'}, inplace=True)
            result.loc[len(result)] = {'Site': 'Total', 'Parents Served (Total)': result['Parents Served (Total)'].sum()}

            # ── Demographics / Missing ────────────────────────────────────────
            st.write("🌸 Processing Participant Demographics…")
            df_part_demo = students_sheets['Participant Demographics'].copy()
            df_part_demo.columns = df_part_demo.iloc[2]
            df_part_demo = df_part_demo.iloc[3:].reset_index(drop=True)

            def summarize_missing_by_school(df, columns_to_check, category_col='Site'):
                if category_col not in df.columns:
                    raise ValueError(f'{category_col} not found in columns')

                missing_site_rows = df[df[category_col].isna() | (df[category_col].astype(str).str.strip() == '')].copy()

                subset = df[columns_to_check + [category_col]].copy()
                subset_fg = subset[subset[category_col].notna()].copy()
                subset_fg[category_col] = subset_fg[category_col].astype(str).str.title()

                for col in columns_to_check:
                    cleaned = subset_fg[col].astype(str).str.strip()
                    nem = cleaned.str.lower() == 'not entered'
                    if col == 'Gender':
                        vg = cleaned.str.title().isin(['Male', 'Female', 'Non-Binary'])
                        subset_fg[col + '_missing'] = ((~vg) | nem).astype(int)
                    else:
                        subset_fg[col + '_missing'] = (subset_fg[col].isna() | nem).astype(int)

                pid = df.loc[subset_fg.index, 'ParticipantID'].astype(str).str.strip()
                spid = df.loc[subset_fg.index, 'State ParticipantID'].astype(str).str.strip()
                vp = pid.str.match(r'^[12]\d{8}$')
                vs = spid.str.match(r'^\d{10}$')
                subset_fg['ParticipantID_missing'] = (~vp).astype(int)
                subset_fg['State ParticipantID_missing'] = (~vs).astype(int)

                missing_cols = (
                    [col + '_missing' for col in columns_to_check]
                    + ['ParticipantID_missing', 'State ParticipantID_missing']
                )

                pivot = subset_fg.groupby(category_col)[missing_cols].sum().reset_index()
                total_r = pd.DataFrame(pivot[missing_cols].sum()).T
                total_r[category_col] = 'Total'
                pivot = pd.concat([pivot, total_r], ignore_index=True)

                all_mf = df.copy()
                pid_a = all_mf['ParticipantID'].astype(str).str.strip()
                spid_a = all_mf['State ParticipantID'].astype(str).str.strip()
                vpa = pid_a.str.match(r'^[12]\d{8}$')
                vsa = spid_a.str.match(r'^\d{10}$')

                for col in columns_to_check:
                    cleaned = all_mf[col].astype(str).str.strip()
                    nem = cleaned.str.lower() == 'not entered'
                    if col == 'Gender':
                        vg = cleaned.str.title().isin(['Male', 'Female', 'Non-Binary'])
                        all_mf[col + '_missing'] = ((~vg) | nem).astype(int)
                    else:
                        all_mf[col + '_missing'] = (all_mf[col].isna() | nem).astype(int)

                all_mf['ParticipantID_missing'] = (~vpa).astype(int)
                all_mf['State ParticipantID_missing'] = (~vsa).astype(int)

                dob_parsed = pd.to_datetime(all_mf['Date Of Birth'], errors='coerce')
                all_mf['DOB_too_young'] = ((dob_parsed.dt.year > 2023) | (dob_parsed.dt.year < 2004)).astype(int)

                flag_cols2 = [col + '_missing' for col in columns_to_check] + ['ParticipantID_missing', 'DOB_too_young']
                total_missing_rows = all_mf[all_mf[flag_cols2].sum(axis=1) > 0].copy()
                young_dob_rows = all_mf[all_mf['DOB_too_young'] == 1].copy()

                dob_young_counts = all_mf.groupby(category_col)['DOB_too_young'].sum()
                pivot = pivot.set_index('Site')
                pivot['Date Of Birth_missing'] += pivot.index.map(dob_young_counts).fillna(0).astype(int)
                pivot = pivot.reset_index()

                pivot = pivot.rename(columns={
                    'Date Of Birth_missing': 'DOB_missing',
                    'State ParticipantID_missing': '10digit_State ParticipantID_missing'
                })[['Site', 'DOB_missing', 'ParticipantID_missing', 'Grade Level_missing', 'Gender_missing', '10digit_State ParticipantID_missing']]

                return pivot, missing_site_rows, total_missing_rows, flag_cols2, young_dob_rows

            columns_of_interest = ['Date Of Birth', 'Grade Level', 'Gender']
            missing_summary, missing_site_rows, total_missing_rows, flag_cols, young_dob_rows = \
                summarize_missing_by_school(df_part_demo, columns_of_interest)

            # ── Staff Missing Info ────────────────────────────────────────────
            # Strictly follows the notebook's `summarize_staff_missing_info` logic.
            st.write("👥 Processing Staff Details…")
            df_staff = pd.read_excel(io.BytesIO(all_bytes), sheet_name="Staff Details", skiprows=2)

            def summarize_staff_missing_info(df_staff, site_col="Site"):
                df = df_staff.copy()

                # Keep rows with a valid site
                df = df[df[site_col].notna() & (df[site_col].astype(str).str.strip() != "")].copy()

                # Make sure key columns exist
                for col in ["Email Address", "First Name", "Last Name", "Staff Type", "Compensation Type", "Funder"]:
                    if col not in df.columns:
                        df[col] = ""

                # Create a unique staff ID. Email is best; if email is blank, use first + last name.
                email = df["Email Address"].astype(str).str.strip().str.lower()
                name_id = (
                    df["First Name"].astype(str).str.strip().str.lower()
                    + "|"
                    + df["Last Name"].astype(str).str.strip().str.lower()
                )
                df["_staff_id"] = np.where(
                    email.ne("") & email.ne("nan"),
                    email,
                    name_id
                )

                # Count each staff member once per site
                df = df.drop_duplicates(subset=[site_col, "_staff_id"]).copy()

                staff_type = df["Staff Type"].astype(str).str.strip()
                comp_type = df["Compensation Type"].astype(str).str.strip()
                funder = df["Funder"].astype(str).str.strip()

                # Staff Type: blank / Not Entered / Other are not accepted
                df["Staff Type_missing"] = (
                    staff_type.eq("")
                    | staff_type.str.lower().isin(["nan", "not entered", "other"])
                ).astype(int)

                # Compensation Type: must be Paid or Volunteer
                df["Employment Type_missing"] = (
                    comp_type.eq("")
                    | comp_type.str.lower().isin(["nan", "not entered"])
                    | ~comp_type.str.lower().isin(["paid", "volunteer"])
                ).astype(int)

                # Funder:
                # - Paid staff must have a funder and it should be 21st CCLC.
                # - Volunteers may leave Funder blank.
                is_volunteer = comp_type.str.lower().eq("volunteer")
                funder_blank = funder.eq("") | funder.str.lower().isin(["nan", "not entered"])
                funder_not_21cclc = ~funder.str.contains("21", case=False, na=False)

                df["Funded by 21st CCLC_missing"] = (
                    ((~is_volunteer) & funder_blank)
                    | ((~funder_blank) & funder_not_21cclc)
                ).astype(int)

                staff_missing_summary = (
                    df.groupby(site_col)
                    .agg(
                        **{
                            "# of Active Program Staff": ("_staff_id", "nunique"),
                            "Staff Type": ("Staff Type_missing", "sum"),
                            "Employment Type": ("Employment Type_missing", "sum"),
                            "Funded by 21st CCLC": ("Funded by 21st CCLC_missing", "sum"),
                        }
                    )
                    .reset_index()
                    .rename(columns={site_col: "Site"})
                )

                total_row = {"Site": "Total"}
                for col in staff_missing_summary.columns:
                    if col != "Site":
                        total_row[col] = staff_missing_summary[col].sum()

                staff_missing_summary = pd.concat(
                    [staff_missing_summary, pd.DataFrame([total_row])],
                    ignore_index=True
                )

                staff_flag_cols = [
                    "Staff Type_missing",
                    "Employment Type_missing",
                    "Funded by 21st CCLC_missing",
                ]

                staff_missing_rows = df[df[staff_flag_cols].sum(axis=1) > 0].copy()

                return staff_missing_summary, staff_missing_rows, staff_flag_cols

            staff_missing_summary, staff_missing_rows, staff_flag_cols = summarize_staff_missing_info(df_staff)

            # ── Site Summary Report ───────────────────────────────────────────
            st.write("🪻 Building Site Summary Report…")
            all_io = io.BytesIO(all_bytes)
            df_act = pd.read_excel(all_io, sheet_name="Activity-Session Details", skiprows=2)
            all_io.seek(0)
            df_enr = pd.read_excel(all_io, sheet_name="Session Enrollment by Session", skiprows=2)
            all_io.seek(0)
            # Excel limits sheet names to 31 chars, so 'Summary' is cut off to 'Summa'
            df_att = pd.read_excel(all_io, sheet_name="Daily Activity Attendance Summa", skiprows=4)

            cols_act = ['Site', 'Activity', 'Session', 'Days Scheduled', 'Session Start Date']
            df_act = df_act[cols_act].copy()
            df_act['Days Scheduled'] = pd.to_numeric(df_act['Days Scheduled'], errors='coerce')
            df_act['Session Start Date'] = pd.to_datetime(df_act['Session Start Date'], errors='coerce')
            today = pd.Timestamp.today().normalize()

            df_enr = df_enr[['Site', 'Activity', 'Session', 'Enrolled Count']].copy()
            df_enr['Enrolled Count'] = pd.to_numeric(df_enr['Enrolled Count'], errors='coerce')
            df_enr.rename(columns={'Enrolled Count': 'Enrolled Participant'}, inplace=True)

            def extract_average(val):
                if pd.isna(val):
                    return np.nan
                try:
                    return float(str(val).replace('Average:', '').strip())
                except:
                    return np.nan

            df_att = df_att[['Site', 'Activity', 'Session', 'Total']].copy()
            df_att['Total'] = df_att['Total'].apply(extract_average).round(0)
            df_att.rename(columns={'Total': 'Average Daily Attendance'}, inplace=True)

            sites = [
                s for s in df_act['Site'].dropna().unique()
                if str(s).strip() != '' and not str(s).startswith('Total') and not str(s).startswith('Average')
            ]

            site_tables = {}
            for site in sites:
                m = pd.merge(
                    df_act[df_act['Site'] == site],
                    df_enr[df_enr['Site'] == site],
                    on=['Site', 'Activity', 'Session'],
                    how='outer'
                )
                m = pd.merge(
                    m,
                    df_att[df_att['Site'] == site],
                    on=['Site', 'Activity', 'Session'],
                    how='outer'
                )
                m = m[~(m['Session Start Date'] >= today)].drop(columns=['Session Start Date'], errors='ignore')
                m = m.fillna("-")
                m = m.sort_values(['Session']).reset_index(drop=True)
                site_tables[site] = m

            # ── Write Excel ───────────────────────────────────────────────────
            st.write("🌈 Writing Excel report…")
            hide_cols = {'Race/Ethnicity', 'English Learner Status', 'Lunch Status', 'Special Education Status', 'IDEA Disability Type'}
            output_buffer = io.BytesIO()

            with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                df_totals.to_excel(writer, sheet_name='Student Summary Statistics', index=False)
                result.to_excel(writer, sheet_name='Family Component Summary', index=False)
                missing_summary.to_excel(writer, sheet_name='Missing Student Summary', index=False)
                missing_site_rows.to_excel(writer, sheet_name="Pull out - Missing Site Info", index=False)

                display_cols = [
                    c for c in total_missing_rows.columns
                    if not c.endswith('_missing') and c != 'DOB_too_young' and c not in hide_cols
                ]

                for site, group in total_missing_rows.groupby('Site'):
                    safe_name = clean_sheet_name('Missing - ' + str(site), fallback='Missing Site')
                    group[display_cols].to_excel(writer, sheet_name=safe_name, index=False)

                young_dc = [c for c in young_dob_rows.columns if not c.endswith('_missing') and c != 'DOB_too_young']
                young_dob_rows[young_dc].to_excel(writer, sheet_name='Pull out - Young DOB', index=False)

                # Staff missing summary and pull-out
                staff_missing_summary.to_excel(writer, sheet_name='Missing Staff Summary', index=False)

                staff_display_cols = [
                    c for c in staff_missing_rows.columns
                    if not c.endswith('_missing') and c != '_staff_id'
                ]
                staff_missing_rows[staff_display_cols].to_excel(writer, sheet_name='Pull out - Missing Staff Info', index=False)

                for site_name, final_df in site_tables.items():
                    safe = clean_sheet_name(site_name, fallback='Site')
                    final_df.to_excel(writer, sheet_name=safe, index=False)

            # ── Apply highlights ──────────────────────────────────────────────
            output_buffer.seek(0)
            wb = load_workbook(output_buffer)

            red_fill = PatternFill('solid', start_color='FF9999', end_color='FF9999')
            blue_fill = PatternFill('solid', start_color='9999FF', end_color='9999FF')

            flag_to_original = {
                fc: fc[:-len('_missing')]
                for fc in flag_cols
                if fc.endswith('_missing') and fc != 'State ParticipantID_missing'
            }

            for site, group in total_missing_rows.groupby('Site'):
                sheet_name = clean_sheet_name('Missing - ' + str(site), fallback='Missing Site')
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                header = {cell.value: cell.column for cell in ws[1]}
                for row_idx, (_, row) in enumerate(group.iterrows(), start=2):
                    for flag_col, orig_col in flag_to_original.items():
                        if orig_col in header and row.get(flag_col, 0) == 1:
                            ws.cell(row=row_idx, column=header[orig_col]).fill = red_fill

            # Staff missing info — red highlights
            if 'Pull out - Missing Staff Info' in wb.sheetnames:
                ws_staff = wb['Pull out - Missing Staff Info']
                staff_header = {cell.value: cell.column for cell in ws_staff[1]}

                # Map each flag column to the original column shown in the pull-out sheet.
                staff_flag_to_original = {
                    'Staff Type_missing': 'Staff Type',
                    'Employment Type_missing': 'Compensation Type',
                    'Funded by 21st CCLC_missing': 'Funder',
                }

                for row_idx, (_, row) in enumerate(staff_missing_rows.iterrows(), start=2):
                    for flag_col in staff_flag_cols:
                        orig_col = staff_flag_to_original.get(flag_col, flag_col.replace('_missing', ''))
                        if orig_col in staff_header and row.get(flag_col, 0) == 1:
                            ws_staff.cell(row=row_idx, column=staff_header[orig_col]).fill = red_fill

            # Young DOB — blue highlights
            if 'Pull out - Young DOB' in wb.sheetnames:
                ws2 = wb['Pull out - Young DOB']
                header2 = {cell.value: cell.column for cell in ws2[1]}
                if 'Date Of Birth' in header2:
                    dob_col_idx = header2['Date Of Birth']
                    for row_idx in range(2, len(young_dob_rows) + 2):
                        ws2.cell(row=row_idx, column=dob_col_idx).fill = blue_fill

            # ── Highlight missing cells in site summary sheets ────────────────
            # New behavior: highlight ANY blank/None/"-" cell across the sheet
            for site_name in site_tables:
                safe = clean_sheet_name(site_name, fallback='Site')
                if safe not in wb.sheetnames:
                    continue
                ws_site = wb[safe]
                for row_idx in range(2, ws_site.max_row + 1):
                    for col_idx in range(1, ws_site.max_column + 1):
                        cell = ws_site.cell(row=row_idx, column=col_idx)
                        if cell.value is None or str(cell.value).strip() in ['', '-']:
                            cell.fill = red_fill

            # ── Copy raw sheets from source files ─────────────────────────────
            st.write("📋 Copying source sheets…")

            def copy_sheet(src_ws, dest_wb, dest_name, skip_rows=0):
                dest_ws = dest_wb.create_sheet(title=dest_name)
                for row in src_ws.iter_rows(min_row=skip_rows + 1):
                    for cell in row:
                        dest_row = cell.row - skip_rows
                        dest_cell = dest_ws.cell(row=dest_row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            dest_cell.font = copy(cell.font)
                            dest_cell.border = copy(cell.border)
                            dest_cell.fill = copy(cell.fill)
                            dest_cell.number_format = cell.number_format
                            dest_cell.alignment = copy(cell.alignment)

                for key, dim in src_ws.column_dimensions.items():
                    dest_ws.column_dimensions[key].width = dim.width
                    dest_ws.column_dimensions[key].hidden = dim.hidden

                for key, dim in src_ws.row_dimensions.items():
                    if dim.index > skip_rows:
                        dest_ws.row_dimensions[dim.index - skip_rows].height = dim.height
                        dest_ws.row_dimensions[dim.index - skip_rows].hidden = dim.hidden

            students_wb_src = load_workbook(io.BytesIO(students_bytes))
            adults_wb_src = load_workbook(io.BytesIO(adults_bytes))

            # Locate source sheets by name (more robust than indices)
            def find_sheet(wb_src, keyword):
                for name in wb_src.sheetnames:
                    if keyword.lower() in name.lower():
                        return wb_src[name]
                return None

            students_phb = find_sheet(students_wb_src, 'Participants By Hour Band')
            adults_phb = find_sheet(adults_wb_src, 'Participants By Hour Band')
            students_demo = find_sheet(students_wb_src, 'Participant Demographics')

            if students_phb is not None:
                copy_sheet(students_phb, wb, 'Students - Participants By Hour', skip_rows=4)
            if adults_phb is not None:
                copy_sheet(adults_phb, wb, 'Adults - Participants By Hour', skip_rows=4)
            if students_demo is not None:
                copy_sheet(students_demo, wb, 'Participant Demographics', skip_rows=3)

            # ── Apply red/green color to the 15+ % column on Student Summary ──
            # Only the "15+ hrs (% of Target)" column gets colored: red <100%, green >=100%.
            # Colors are baked into Arial Narrow 10pt fonts so the workbook-wide font
            # pass below doesn't overwrite them.
            arial_narrow_default = Font(name='Arial Narrow', size=10)
            arial_narrow_red = Font(name='Arial Narrow', size=10, color='FF0000')
            arial_narrow_green = Font(name='Arial Narrow', size=10, color='008000')

            # Track which (sheet, row, col) cells should keep a colored font
            colored_cells = {}  # (sheet_name, row, col) -> Font

            if 'Student Summary Statistics' in wb.sheetnames:
                ws_sss = wb['Student Summary Statistics']
                header_map = {cell.value: cell.column for cell in ws_sss[1]}
                target_col = '# of students 15+ hrs total (% of Target)'
                if target_col in header_map:
                    col_idx = header_map[target_col]
                    for row_idx in range(2, ws_sss.max_row + 1):
                        cell = ws_sss.cell(row=row_idx, column=col_idx)
                        try:
                            pct = int(str(cell.value).split('(')[1].replace('%)', '').strip())
                            colored_cells[('Student Summary Statistics', row_idx, col_idx)] = (
                                arial_narrow_green if pct >= 100 else arial_narrow_red
                            )
                        except Exception:
                            pass

            # ── Apply Arial Narrow 10pt font to all cells ─────────────────────
            # Cells flagged in colored_cells get a colored Arial Narrow font instead.
            for ws_name in wb.sheetnames:
                for row in wb[ws_name].iter_rows():
                    for cell in row:
                        key = (ws_name, cell.row, cell.column)
                        cell.font = colored_cells.get(key, arial_narrow_default)

            final_buffer = io.BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)

            status.update(label="✅ Report ready!", state="complete")
            st.session_state.output_bytes = final_buffer.read()
            st.session_state.report_ready = True

        except Exception as e:
            status.update(label="❌ Something went wrong", state="error")
            st.error(f"Error: {e}")
            st.session_state.report_ready = False

if st.session_state.get("report_ready"):
    safe_institution = institution_name.strip().replace(' ', '_') or 'Institution'
    today_str = date.today().strftime('%Y%m%d')
    output_filename = f"{today_str}_{safe_institution}_WeeklyUpdates.xlsx"
    st.success("🎉 Your report is ready!")
    st.download_button(
        label=f"⬇️ Download {output_filename}",
        data=st.session_state.output_bytes,
        file_name=output_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )
