"""
CSD3 Weekly Report Processor — Streamlit App
Upload 3 source Excel files, set targets, get a fully-formatted Excel report.
"""

import io
import os
import tempfile

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ----------------------------------------------------------------------
# Core processing functions (ported from notebook)
# ----------------------------------------------------------------------

def summarize_missing_by_school(df, columns_to_check, category_col="Site"):
    if category_col not in df.columns:
        raise ValueError(f"{category_col} not found in DataFrame columns")

    missing_site_rows = df[
        df[category_col].isna() | (df[category_col].astype(str).str.strip() == "")
    ].copy()

    subset = df[columns_to_check + [category_col]].copy()
    subset_for_grouping = subset[subset[category_col].notna()].copy()
    subset_for_grouping[category_col] = (
        subset_for_grouping[category_col].astype(str).str.title()
    )

    for col in columns_to_check:
        cleaned = subset_for_grouping[col].astype(str).str.strip()
        not_entered_mask = cleaned.str.lower() == "not entered"
        if col == "Gender":
            valid_gender = cleaned.str.title().isin(["Male", "Female", "Non-Binary"])
            subset_for_grouping[col + "_missing"] = (
                (~valid_gender) | not_entered_mask
            ).astype(int)
        else:
            subset_for_grouping[col + "_missing"] = (
                subset_for_grouping[col].isna() | not_entered_mask
            ).astype(int)

    pid = df.loc[subset_for_grouping.index, "ParticipantID"].astype(str).str.strip()
    spid = df.loc[subset_for_grouping.index, "State ParticipantID"].astype(str).str.strip()
    valid_pid = pid.str.match(r"^[12]\d{8}$")
    valid_spid = spid.str.match(r"^\d{10}$")
    subset_for_grouping["ParticipantID_missing"] = (~valid_pid).astype(int)
    subset_for_grouping["State ParticipantID_missing"] = (~valid_spid).astype(int)

    missing_cols = [col + "_missing" for col in columns_to_check] + [
        "ParticipantID_missing",
        "State ParticipantID_missing",
    ]
    pivot = (
        subset_for_grouping.groupby(category_col)[missing_cols].sum().reset_index()
    )
    total_row = pd.DataFrame(pivot[missing_cols].sum()).T
    total_row[category_col] = "Total"
    pivot = pd.concat([pivot, total_row], ignore_index=True)

    all_missing_flags = df.copy()
    pid_all = all_missing_flags["ParticipantID"].astype(str).str.strip()
    spid_all = all_missing_flags["State ParticipantID"].astype(str).str.strip()
    valid_pid_all = pid_all.str.match(r"^[12]\d{8}$")
    valid_spid_all = spid_all.str.match(r"^\d{10}$")

    for col in columns_to_check:
        cleaned = all_missing_flags[col].astype(str).str.strip()
        not_entered_mask = cleaned.str.lower() == "not entered"
        if col == "Gender":
            valid_gender = cleaned.str.title().isin(["Male", "Female", "Non-Binary"])
            all_missing_flags[col + "_missing"] = (
                (~valid_gender) | not_entered_mask
            ).astype(int)
        else:
            all_missing_flags[col + "_missing"] = (
                all_missing_flags[col].isna() | not_entered_mask
            ).astype(int)

    all_missing_flags["ParticipantID_missing"] = (~valid_pid_all).astype(int)
    all_missing_flags["State ParticipantID_missing"] = (~valid_spid_all).astype(int)

    dob_parsed = pd.to_datetime(all_missing_flags["Date Of Birth"], errors="coerce")
    all_missing_flags["DOB_too_young"] = (
        (dob_parsed.dt.year > 2023) | (dob_parsed.dt.year < 2004)
    ).astype(int)

    flag_cols = [col + "_missing" for col in columns_to_check] + [
        "ParticipantID_missing",
        "DOB_too_young",
    ]
    total_missing_rows = all_missing_flags[
        all_missing_flags[flag_cols].sum(axis=1) > 0
    ].copy()

    young_dob_rows = all_missing_flags[all_missing_flags["DOB_too_young"] == 1].copy()

    dob_young_counts = all_missing_flags.groupby(category_col)["DOB_too_young"].sum()
    pivot = pivot.set_index("Site")
    pivot["Date Of Birth_missing"] += (
        pivot.index.map(dob_young_counts).fillna(0).astype(int)
    )
    pivot = pivot.reset_index()

    pivot = pivot.rename(
        columns={
            "Date Of Birth_missing": "DOB_missing",
            "State ParticipantID_missing": "10digit_State ParticipantID_missing",
        }
    )[
        [
            "Site",
            "DOB_missing",
            "ParticipantID_missing",
            "Grade Level_missing",
            "Gender_missing",
            "10digit_State ParticipantID_missing",
        ]
    ]

    return pivot, missing_site_rows, total_missing_rows, flag_cols, young_dob_rows


def summarize_staff_missing_info(df_staff, site_col="Site"):
    df = df_staff.copy()
    df = df[
        df[site_col].notna() & (df[site_col].astype(str).str.strip() != "")
    ].copy()

    for col in [
        "Email Address",
        "First Name",
        "Last Name",
        "Staff Type",
        "Compensation Type",
        "Funder",
    ]:
        if col not in df.columns:
            df[col] = ""

    email = df["Email Address"].astype(str).str.strip().str.lower()
    name_id = (
        df["First Name"].astype(str).str.strip().str.lower()
        + "|"
        + df["Last Name"].astype(str).str.strip().str.lower()
    )
    df["_staff_id"] = np.where(
        email.ne("") & email.ne("nan"), email, name_id
    )

    df = df.drop_duplicates(subset=[site_col, "_staff_id"]).copy()

    staff_type = df["Staff Type"].astype(str).str.strip()
    comp_type = df["Compensation Type"].astype(str).str.strip()
    funder = df["Funder"].astype(str).str.strip()

    df["Staff Type_missing"] = (
        staff_type.eq("")
        | staff_type.str.lower().isin(["nan", "not entered", "other"])
    ).astype(int)

    df["Employment Type_missing"] = (
        comp_type.eq("")
        | comp_type.str.lower().isin(["nan", "not entered"])
        | ~comp_type.str.lower().isin(["paid", "volunteer"])
    ).astype(int)

    is_volunteer = comp_type.str.lower().eq("volunteer")
    funder_blank = funder.eq("") | funder.str.lower().isin(["nan", "not entered"])
    funder_not_21cclc = ~funder.str.contains("21", case=False, na=False)

    df["Funded by 21st CCLC_missing"] = (
        ((~is_volunteer) & funder_blank)
        | ((~funder_blank) & funder_not_21cclc)
    ).astype(int)

    staff_missing_summary = (
        df.groupby(site_col)
        .agg(
            **{
                "# of Active Program Staff": ("_staff_id", "nunique"),
                "Staff Type": ("Staff Type_missing", "sum"),
                "Employment Type": ("Employment Type_missing", "sum"),
                "Funded by 21st CCLC": ("Funded by 21st CCLC_missing", "sum"),
            }
        )
        .reset_index()
        .rename(columns={site_col: "Site"})
    )

    total_row = {"Site": "Total"}
    for col in staff_missing_summary.columns:
        if col != "Site":
            total_row[col] = staff_missing_summary[col].sum()

    staff_missing_summary = pd.concat(
        [staff_missing_summary, pd.DataFrame([total_row])], ignore_index=True
    )

    staff_flag_cols = [
        "Staff Type_missing",
        "Employment Type_missing",
        "Funded by 21st CCLC_missing",
    ]

    staff_missing_rows = df[df[staff_flag_cols].sum(axis=1) > 0].copy()

    return staff_missing_summary, staff_missing_rows, staff_flag_cols


def apply_missing_highlights(
    output_filename,
    total_missing_rows,
    flag_cols,
    young_dob_rows,
    staff_missing_rows=None,
    staff_flag_cols=None,
    site_tables=None,
):
    red_fill = PatternFill("solid", start_color="FF9999", end_color="FF9999")
    blue_fill = PatternFill("solid", start_color="9999FF", end_color="9999FF")

    flag_to_original = {
        fc: fc[: -len("_missing")]
        for fc in flag_cols
        if fc.endswith("_missing") and fc != "State ParticipantID_missing"
    }

    wb = load_workbook(output_filename)

    for site, group in total_missing_rows.groupby("Site"):
        sheet_name = ("Missing - " + str(site))[:31]
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = {cell.value: cell.column for cell in ws[1]}
        for row_idx, (_, row) in enumerate(group.iterrows(), start=2):
            for flag_col, orig_col in flag_to_original.items():
                if orig_col in header and row.get(flag_col, 0) == 1:
                    ws.cell(row=row_idx, column=header[orig_col]).fill = red_fill

    if (
        staff_missing_rows is not None
        and staff_flag_cols is not None
        and "Pull out - Missing Staff Info" in wb.sheetnames
    ):
        ws_staff = wb["Pull out - Missing Staff Info"]
        staff_header = {cell.value: cell.column for cell in ws_staff[1]}
        staff_flag_to_original = {
            "Staff Type_missing": "Staff Type",
            "Employment Type_missing": "Compensation Type",
            "Funded by 21st CCLC_missing": "Funder",
        }
        for row_idx, (_, row) in enumerate(staff_missing_rows.iterrows(), start=2):
            for flag_col in staff_flag_cols:
                orig_col = staff_flag_to_original.get(
                    flag_col, flag_col.replace("_missing", "")
                )
                if orig_col in staff_header and row.get(flag_col, 0) == 1:
                    ws_staff.cell(
                        row=row_idx, column=staff_header[orig_col]
                    ).fill = red_fill

    if "Pull out - Young DOB" in wb.sheetnames:
        ws2 = wb["Pull out - Young DOB"]
        header2 = {cell.value: cell.column for cell in ws2[1]}
        if "Date Of Birth" in header2:
            dob_col_idx = header2["Date Of Birth"]
            for row_idx in range(2, len(young_dob_rows) + 2):
                ws2.cell(row=row_idx, column=dob_col_idx).fill = blue_fill

    if site_tables is not None:
        for site_name, final_df in site_tables.items():
            safe_sheet_name = (
                str(site_name)[:31]
                .replace(":", "")
                .replace("/", "")
                .replace("\\", "")
                .replace("?", "")
                .replace("*", "")
            )
            if safe_sheet_name not in wb.sheetnames:
                continue
            ws_site = wb[safe_sheet_name]
            for row_idx in range(2, ws_site.max_row + 1):
                for col_idx in range(1, ws_site.max_column + 1):
                    cell = ws_site.cell(row=row_idx, column=col_idx)
                    if cell.value is None or str(cell.value).strip() in ["", "-"]:
                        cell.fill = red_fill

    wb.save(output_filename)


def color_pct_cols(val):
    try:
        pct = int(val.split("(")[1].replace("%)", "").strip())
        return "color: green" if pct >= 100 else "color: red"
    except Exception:
        return ""


def process_reports(students_file, adults_file, all_file, target_values):
    """Run the full pipeline. Returns bytes of the output Excel file."""

    students_sheets = pd.read_excel(students_file, sheet_name=None)

    # ---- Student Summary Statistics ----
    df_part_by_hour = students_sheets["Participants By Hour Band (Site"]
    df_part_by_hour.columns = df_part_by_hour.iloc[1]
    df_part_by_hour = df_part_by_hour.iloc[2:].reset_index(drop=True)

    rows_to_drop = []
    for idx, row in df_part_by_hour.iterrows():
        institution = row["Institution"]
        site = row["Site"]
        is_empty_institution = (
            pd.isna(institution) or str(institution).strip() == "NaN"
        )
        is_empty_site = pd.isna(site) or str(site).strip() == "NaN"
        if (is_empty_institution and is_empty_site) or not is_empty_institution:
            rows_to_drop.append(idx)
    df_part_by_hour = df_part_by_hour.drop(index=rows_to_drop).reset_index(drop=True)

    daily_site_att = students_sheets["Daily Site Attendance Summary"]
    daily_site_att.columns = daily_site_att.iloc[2]
    daily_site_att = daily_site_att.iloc[3:]
    daily_site_att.columns.name = None
    daily_site_att = daily_site_att.reset_index(drop=True)
    daily_site_att = daily_site_att[["Total"]]
    daily_site_att = daily_site_att.iloc[:-1]
    daily_site_att["Total"] = daily_site_att["Total"].str.extract(r"(\d+\.?\d*)")

    all_cols = [
        "0 Hours",
        "Less Than 15 Hours",
        "15-44 Hours",
        "45-89 Hours",
        "90-179 Hours",
        "180-269 Hours",
        "270+ Hours",
    ]
    served_cols = [
        "Less Than 15 Hours",
        "15-44 Hours",
        "45-89 Hours",
        "90-179 Hours",
        "180-269 Hours",
        "270+ Hours",
    ]
    plus15_cols = [
        "15-44 Hours",
        "45-89 Hours",
        "90-179 Hours",
        "180-269 Hours",
        "270+ Hours",
    ]
    plus90_cols = ["90-179 Hours", "180-269 Hours", "270+ Hours"]

    existing_cols = df_part_by_hour.columns.tolist()
    all_cols = [c for c in all_cols if c in existing_cols]
    served_cols = [c for c in served_cols if c in existing_cols]
    plus15_cols = [c for c in plus15_cols if c in existing_cols]
    plus90_cols = [c for c in plus90_cols if c in existing_cols]

    df_part_by_hour[all_cols] = (
        df_part_by_hour[all_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
    )

    df_totals = pd.DataFrame(
        {
            "[Total # Enrolled]": df_part_by_hour[all_cols].sum(axis=1),
            "[Total # Served]": df_part_by_hour[served_cols].sum(axis=1),
            "[Total # 15+]": df_part_by_hour[plus15_cols].sum(axis=1),
            "[Total # 90+]": df_part_by_hour[plus90_cols].sum(axis=1),
        }
    )

    if len(target_values) != len(df_totals):
        raise ValueError(
            f"Number of target values ({len(target_values)}) does not match "
            f"number of sites ({len(df_totals)})."
        )

    df_totals.insert(0, "[Target # of students served]", target_values)

    daily_site_att["Total"] = daily_site_att["Total"].astype(int)
    daily_site_att = daily_site_att.rename(
        columns={"Total": "Avg. # of Students Per Day"}
    )
    df_totals.insert(
        3,
        "Avg. # of Students Per Day",
        daily_site_att["Avg. # of Students Per Day"].values,
    )

    school_names = [
        row["Site"]
        for _, row in df_part_by_hour.iterrows()
        if pd.notna(row["Site"])
        and row["Site"] != "Subtotal"
        and row["Site"] != "Total"
    ]
    df_totals.insert(0, "School", school_names)

    total_row = pd.DataFrame(df_totals.iloc[:, 1:].sum()).T
    total_row.insert(0, "School", "Total")
    df_totals = pd.concat([df_totals, total_row], ignore_index=True)

    df_totals["# of students 15+ hrs total (% of Target)"] = (
        df_totals["[Total # 15+]"].astype(int).astype(str)
        + " ("
        + (
            (df_totals["[Total # 15+]"] / df_totals["[Target # of students served]"])
            * 100
        )
        .round()
        .astype(int)
        .astype(str)
        + "%)"
    )
    df_totals["# of students 90+ hrs total (% of Target)"] = (
        df_totals["[Total # 90+]"].astype(int).astype(str)
        + " ("
        + (
            (df_totals["[Total # 90+]"] / df_totals["[Target # of students served]"])
            * 100
        )
        .round()
        .astype(int)
        .astype(str)
        + "%)"
    )
    df_totals = df_totals.drop(columns=["[Total # 15+]", "[Total # 90+]"]).reset_index(
        drop=True
    )

    styled_totals = df_totals.style.map(
        color_pct_cols, subset=["# of students 15+ hrs total (% of Target)"]
    )

    # ---- Family Component ----
    df_hours = pd.read_excel(
        adults_file, sheet_name="Participant Attendance Hours", skiprows=2
    )
    df_hours["HoursPresent"] = pd.to_numeric(df_hours["HoursPresent"], errors="coerce")
    df_hours["ParticipantId"] = (
        df_hours["ParticipantId"].astype(str).str.replace(r"\.0$", "", regex=True)
    )
    df_active = df_hours[
        (df_hours["HoursPresent"] > 0) & (df_hours["ParticipantId"].str.len() != 9)
    ]
    parents_result = (
        df_active.groupby("Site")["ParticipantId"].nunique().reset_index()
    )
    parents_result.rename(
        columns={"ParticipantId": "Parents Served (Total)"}, inplace=True
    )
    parents_result.loc[len(parents_result)] = {
        "Site": "Total",
        "Parents Served (Total)": parents_result["Parents Served (Total)"].sum(),
    }

    # ---- Participant Demographics ----
    df_part_demo = students_sheets["Participant Demographics"]
    df_part_demo.columns = df_part_demo.iloc[2]
    df_part_demo = df_part_demo.iloc[3:].reset_index(drop=True)

    columns_of_interest = ["Date Of Birth", "Grade Level", "Gender"]
    (
        missing_summary,
        missing_site_rows,
        total_missing_rows,
        flag_cols,
        young_dob_rows,
    ) = summarize_missing_by_school(df_part_demo, columns_of_interest)

    # ---- Staff ----
    df_staff = pd.read_excel(all_file, sheet_name="Staff Details", skiprows=2)
    (
        staff_missing_summary,
        staff_missing_rows,
        staff_flag_cols,
    ) = summarize_staff_missing_info(df_staff)

    # ---- Site Summary Reports ----
    df_act = pd.read_excel(
        all_file, sheet_name="Activity-Session Details", skiprows=2
    )
    df_enr = pd.read_excel(
        all_file, sheet_name="Session Enrollment by Session", skiprows=2
    )
    df_att = pd.read_excel(
        all_file, sheet_name="Daily Activity Attendance Summa", skiprows=4
    )

    cols_act = ["Site", "Activity", "Session", "Days Scheduled", "Session Start Date"]
    df_act = df_act[cols_act].copy()
    df_act["Days Scheduled"] = pd.to_numeric(df_act["Days Scheduled"], errors="coerce")
    df_act["Session Start Date"] = pd.to_datetime(
        df_act["Session Start Date"], errors="coerce"
    )
    today = pd.Timestamp.today().normalize()

    cols_enr = ["Site", "Activity", "Session", "Enrolled Count"]
    df_enr = df_enr[cols_enr].copy()
    df_enr["Enrolled Count"] = pd.to_numeric(df_enr["Enrolled Count"], errors="coerce")
    df_enr.rename(columns={"Enrolled Count": "Enrolled Participant"}, inplace=True)

    cols_att = ["Site", "Activity", "Session", "Total"]
    df_att = df_att[cols_att].copy()

    def extract_average(val):
        if pd.isna(val):
            return np.nan
        val_str = str(val).replace("Average:", "").strip()
        try:
            return float(val_str)
        except ValueError:
            return np.nan

    df_att["Total"] = df_att["Total"].apply(extract_average).round(0)
    df_att.rename(columns={"Total": "Average Daily Attendance"}, inplace=True)

    sites = df_act["Site"].dropna().unique().tolist()
    sites = [
        s
        for s in sites
        if str(s).strip() != ""
        and not str(s).startswith("Total")
        and not str(s).startswith("Average")
    ]

    site_tables = {}
    for site in sites:
        site_act = df_act[df_act["Site"] == site].copy()
        site_enr = df_enr[df_enr["Site"] == site].copy()
        site_att = df_att[df_att["Site"] == site].copy()

        merged = pd.merge(
            site_act, site_enr, on=["Site", "Activity", "Session"], how="outer"
        )
        merged = pd.merge(
            merged, site_att, on=["Site", "Activity", "Session"], how="outer"
        )
        merged = merged[~(merged["Session Start Date"] >= today)]
        if "Session Start Date" in merged.columns:
            merged = merged.drop(columns=["Session Start Date"])
        merged = merged.fillna("-")

        site_tables[site] = merged.sort_values(
            by=["Session"], ascending=True
        ).reset_index(drop=True)

    # ---- Write Excel ----
    tmp_dir = tempfile.mkdtemp()
    output_path = os.path.join(tmp_dir, "report.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        styled_totals.to_excel(
            writer, sheet_name="Student Summary Statistics", index=False
        )
        missing_summary.to_excel(
            writer, sheet_name="Missing Student Summary", index=False
        )
        missing_site_rows.to_excel(
            writer, sheet_name="Pull out - Missing Site Info", index=False
        )

        hide_cols = {
            "Race/Ethnicity",
            "English Learner Status",
            "Lunch Status",
            "Special Education Status",
            "IDEA Disability Type",
        }
        display_cols = [
            c
            for c in total_missing_rows.columns
            if not c.endswith("_missing")
            and c != "DOB_too_young"
            and c not in hide_cols
        ]

        for site, group in total_missing_rows.groupby("Site"):
            safe_name = ("Missing - " + str(site))[:31]
            group[display_cols].to_excel(writer, sheet_name=safe_name, index=False)

        young_display_cols = [
            c
            for c in young_dob_rows.columns
            if not c.endswith("_missing") and c != "DOB_too_young"
        ]
        young_dob_rows[young_display_cols].to_excel(
            writer, sheet_name="Pull out - Young DOB", index=False
        )

        staff_missing_summary.to_excel(
            writer, sheet_name="Missing Staff Summary", index=False
        )
        staff_display_cols = [
            c
            for c in staff_missing_rows.columns
            if not c.endswith("_missing") and c != "_staff_id"
        ]
        staff_missing_rows[staff_display_cols].to_excel(
            writer, sheet_name="Pull out - Missing Staff Info", index=False
        )

        for site_name, final_df in site_tables.items():
            safe_sheet_name = (
                str(site_name)[:31]
                .replace(":", "")
                .replace("/", "")
                .replace("\\", "")
                .replace("?", "")
                .replace("*", "")
            )
            final_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

    apply_missing_highlights(
        output_path,
        total_missing_rows,
        flag_cols,
        young_dob_rows,
        staff_missing_rows=staff_missing_rows,
        staff_flag_cols=staff_flag_cols,
        site_tables=site_tables,
    )

    with open(output_path, "rb") as f:
        return f.read(), df_totals, parents_result, missing_summary, staff_missing_summary


# ----------------------------------------------------------------------
# Streamlit UI
# ----------------------------------------------------------------------

st.set_page_config(page_title="CSD3 Weekly Report Processor", layout="wide")
st.title("CSD3 Weekly Report Processor")
st.caption("Upload the 3 source Excel files, set targets, generate the formatted report.")

with st.sidebar:
    st.header("1. Upload source files")
    students_file = st.file_uploader(
        "Students workbook (Students_CSD3_Weekly Reports...)", type=["xlsx", "xls"]
    )
    adults_file = st.file_uploader(
        "Adults workbook (Adults_CSD3_Weekly Reports...)", type=["xlsx", "xls"]
    )
    all_file = st.file_uploader(
        "All workbook (All_CSD3_Weekly Reports...)", type=["xlsx", "xls"]
    )

    st.header("2. Targets")
    st.caption("Target # of students served per site (one per line, in site order).")
    targets_text = st.text_area(
        "Target values",
        value="152\n200\n100",
        height=120,
    )

    st.header("3. Output filename")
    output_name = st.text_input(
        "Filename (no extension)", value="Processed_Site_Tables"
    )

    run_btn = st.button("Generate Report", type="primary", use_container_width=True)

if run_btn:
    if not (students_file and adults_file and all_file):
        st.error("Please upload all 3 Excel files.")
        st.stop()

    try:
        target_values = [int(x.strip()) for x in targets_text.splitlines() if x.strip()]
    except ValueError:
        st.error("Targets must be integers, one per line.")
        st.stop()

    with st.spinner("Processing..."):
        try:
            (
                excel_bytes,
                df_totals,
                parents_result,
                missing_summary,
                staff_missing_summary,
            ) = process_reports(students_file, adults_file, all_file, target_values)
        except Exception as e:
            st.error(f"Processing failed: {e}")
            st.exception(e)
            st.stop()

    st.success("Report generated.")

    safe_name = output_name.strip() or "Processed_Site_Tables"
    if not safe_name.lower().endswith(".xlsx"):
        safe_name += ".xlsx"

    st.download_button(
        label=f"Download {safe_name}",
        data=excel_bytes,
        file_name=safe_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

    with st.expander("Preview: Student Summary Statistics", expanded=True):
        st.dataframe(df_totals, use_container_width=True)
    with st.expander("Preview: Parents Served"):
        st.dataframe(parents_result, use_container_width=True)
    with st.expander("Preview: Missing Student Summary"):
        st.dataframe(missing_summary, use_container_width=True)
    with st.expander("Preview: Missing Staff Summary"):
        st.dataframe(staff_missing_summary, use_container_width=True)
else:
    st.info("Upload files and click **Generate Report** in the sidebar.")
