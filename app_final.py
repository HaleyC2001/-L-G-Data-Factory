import io
from datetime import date
from copy import copy

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def clean_sheet_name(name, fallback="Sheet"):
    name = str(name).strip()
    for ch in [":", "/", "\\", "?", "*", "[", "]"]:
        name = name.replace(ch, "")
    name = name.strip("'")
    if not name or name.lower() == "nan":
        name = fallback
    return name[:31]


def find_sheet_name(xl_or_wb, keywords, fallback_index=None):
    """Return the first sheet whose name contains all keywords; otherwise fallback by index."""
    names = xl_or_wb.sheet_names if hasattr(xl_or_wb, "sheet_names") else xl_or_wb.sheetnames
    lower_names = {name.lower(): name for name in names}
    for lname, original in lower_names.items():
        if all(k.lower() in lname for k in keywords):
            return original
    if fallback_index is not None and fallback_index < len(names):
        return names[fallback_index]
    raise ValueError(f"Could not find sheet containing: {keywords}")


def set_header_from_row(df, header_row_idx):
    df = df.copy()
    df.columns = df.iloc[header_row_idx]
    df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
    df.columns.name = None
    return df


def clean_participants_by_hour(df_raw):
    # Find the row that contains the actual headers.
    header_idx = None
    for i in range(min(8, len(df_raw))):
        vals = [str(v).strip() for v in df_raw.iloc[i].tolist()]
        if "Site" in vals and any("15" in v for v in vals):
            header_idx = i
            break
    if header_idx is None:
        header_idx = 1

    df = set_header_from_row(df_raw, header_idx)

    # Remove rows with institution/header noise if Institution exists.
    if "Institution" in df.columns and "Site" in df.columns:
        rows_to_drop = []
        for idx, row in df.iterrows():
            institution = row.get("Institution")
            site = row.get("Site")
            empty_institution = pd.isna(institution) or str(institution).strip().lower() in ["", "nan"]
            empty_site = pd.isna(site) or str(site).strip().lower() in ["", "nan"]
            if (empty_institution and empty_site) or not empty_institution:
                rows_to_drop.append(idx)
        df = df.drop(index=rows_to_drop).reset_index(drop=True)

    return df


def summarize_missing_by_school(df, columns_to_check, category_col="Site"):
    if category_col not in df.columns:
        raise ValueError(f"{category_col} not found in DataFrame columns")

    missing_site_rows = df[df[category_col].isna() | (df[category_col].astype(str).str.strip() == "")].copy()

    subset = df[columns_to_check + [category_col]].copy()
    subset_for_grouping = subset[subset[category_col].notna()].copy()
    subset_for_grouping[category_col] = subset_for_grouping[category_col].astype(str).str.title()

    for col in columns_to_check:
        cleaned = subset_for_grouping[col].astype(str).str.strip()
        not_entered_mask = cleaned.str.lower() == "not entered"
        if col == "Gender":
            valid_gender = cleaned.str.title().isin(["Male", "Female", "Non-Binary"])
            subset_for_grouping[col + "_missing"] = ((~valid_gender) | not_entered_mask).astype(int)
        else:
            subset_for_grouping[col + "_missing"] = (subset_for_grouping[col].isna() | not_entered_mask).astype(int)

    pid = df.loc[subset_for_grouping.index, "ParticipantID"].astype(str).str.strip()
    spid = df.loc[subset_for_grouping.index, "State ParticipantID"].astype(str).str.strip()
    valid_pid = pid.str.match(r"^[12]\d{8}$")
    valid_spid = spid.str.match(r"^\d{10}$")
    subset_for_grouping["ParticipantID_missing"] = (~valid_pid).astype(int)
    subset_for_grouping["State ParticipantID_missing"] = (~valid_spid).astype(int)

    missing_cols = [col + "_missing" for col in columns_to_check] + ["ParticipantID_missing", "State ParticipantID_missing"]
    pivot = subset_for_grouping.groupby(category_col)[missing_cols].sum().reset_index()
    total_row = pd.DataFrame(pivot[missing_cols].sum()).T
    total_row[category_col] = "Total"
    pivot = pd.concat([pivot, total_row], ignore_index=True)

    all_missing_flags = df.copy()
    pid_all = all_missing_flags["ParticipantID"].astype(str).str.strip()
    spid_all = all_missing_flags["State ParticipantID"].astype(str).str.strip()
    valid_pid_all = pid_all.str.match(r"^[12]\d{8}$")
    valid_spid_all = spid_all.str.match(r"^\d{10}$")

    for col in columns_to_check:
        cleaned = all_missing_flags[col].astype(str).str.strip()
        not_entered_mask = cleaned.str.lower() == "not entered"
        if col == "Gender":
            valid_gender = cleaned.str.title().isin(["Male", "Female", "Non-Binary"])
            all_missing_flags[col + "_missing"] = ((~valid_gender) | not_entered_mask).astype(int)
        else:
            all_missing_flags[col + "_missing"] = (all_missing_flags[col].isna() | not_entered_mask).astype(int)

    all_missing_flags["ParticipantID_missing"] = (~valid_pid_all).astype(int)
    all_missing_flags["State ParticipantID_missing"] = (~valid_spid_all).astype(int)

    dob_parsed = pd.to_datetime(all_missing_flags["Date Of Birth"], errors="coerce")
    all_missing_flags["DOB_too_young"] = ((dob_parsed.dt.year > 2023) | (dob_parsed.dt.year < 2004)).astype(int)

    flag_cols = [col + "_missing" for col in columns_to_check] + ["ParticipantID_missing", "DOB_too_young"]
    total_missing_rows = all_missing_flags[all_missing_flags[flag_cols].sum(axis=1) > 0].copy()
    young_dob_rows = all_missing_flags[all_missing_flags["DOB_too_young"] == 1].copy()

    dob_young_counts = all_missing_flags.groupby(category_col)["DOB_too_young"].sum()
    pivot = pivot.set_index("Site")
    pivot["Date Of Birth_missing"] += pivot.index.map(dob_young_counts).fillna(0).astype(int)
    pivot = pivot.reset_index()

    pivot = pivot.rename(columns={
        "Date Of Birth_missing": "DOB_missing",
        "State ParticipantID_missing": "10digit_State ParticipantID_missing",
    })[["Site", "DOB_missing", "ParticipantID_missing", "Grade Level_missing", "Gender_missing", "10digit_State ParticipantID_missing"]]

    return pivot, missing_site_rows, total_missing_rows, flag_cols, young_dob_rows


def summarize_staff_missing_info(df_staff, site_col="Site"):
    df = df_staff.copy()
    df = df[df[site_col].notna() & (df[site_col].astype(str).str.strip() != "")].copy()

    for col in ["Email Address", "First Name", "Last Name", "Staff Type", "Compensation Type", "Funder"]:
        if col not in df.columns:
            df[col] = ""

    email = df["Email Address"].astype(str).str.strip().str.lower()
    name_id = df["First Name"].astype(str).str.strip().str.lower() + "|" + df["Last Name"].astype(str).str.strip().str.lower()
    df["_staff_id"] = np.where(email.ne("") & email.ne("nan"), email, name_id)

    # Count each staff member once per site.
    df = df.drop_duplicates(subset=[site_col, "_staff_id"]).copy()

    staff_type = df["Staff Type"].astype(str).str.strip()
    comp_type = df["Compensation Type"].astype(str).str.strip()
    funder = df["Funder"].astype(str).str.strip()

    df["Staff Type_missing"] = (
        staff_type.eq("") | staff_type.str.lower().isin(["nan", "not entered", "other"])
    ).astype(int)

    df["Employment Type_missing"] = (
        comp_type.eq("")
        | comp_type.str.lower().isin(["nan", "not entered"])
        | ~comp_type.str.lower().isin(["paid", "volunteer"])
    ).astype(int)

    is_volunteer = comp_type.str.lower().eq("volunteer")
    funder_blank = funder.eq("") | funder.str.lower().isin(["nan", "not entered"])
    funder_not_21cclc = ~funder.str.contains("21", case=False, na=False)

    df["Funded by 21st CCLC_missing"] = (
        ((~is_volunteer) & funder_blank) | ((~funder_blank) & funder_not_21cclc)
    ).astype(int)

    staff_missing_summary = (
        df.groupby(site_col)
        .agg(**{
            "# of Active Program Staff": ("_staff_id", "nunique"),
            "Staff Type": ("Staff Type_missing", "sum"),
            "Employment Type": ("Employment Type_missing", "sum"),
            "Funded by 21st CCLC": ("Funded by 21st CCLC_missing", "sum"),
        })
        .reset_index()
        .rename(columns={site_col: "Site"})
    )

    total_row = {"Site": "Total"}
    for col in staff_missing_summary.columns:
        if col != "Site":
            total_row[col] = staff_missing_summary[col].sum()
    staff_missing_summary = pd.concat([staff_missing_summary, pd.DataFrame([total_row])], ignore_index=True)

    staff_flag_cols = ["Staff Type_missing", "Employment Type_missing", "Funded by 21st CCLC_missing"]
    staff_missing_rows = df[df[staff_flag_cols].sum(axis=1) > 0].copy()
    return staff_missing_summary, staff_missing_rows, staff_flag_cols


def color_pct_cols(val):
    try:
        pct = int(str(val).split("(")[1].replace("%)", "").strip())
        return "color: green" if pct >= 100 else "color: red"
    except Exception:
        return ""


def write_combined_missing_summary(writer, missing_summary, staff_missing_summary):
    """Write student and staff missing summaries side-by-side with visible staff section headers."""
    sheet = "Missing student & staff Summary"
    missing_summary.to_excel(writer, sheet_name=sheet, startrow=1, startcol=0, index=False)
    staff_missing_summary.to_excel(writer, sheet_name=sheet, startrow=1, startcol=8, index=False)

    ws = writer.sheets[sheet]
    blue_fill = PatternFill("solid", start_color="B4C6E7", end_color="B4C6E7")
    thin_gray = Side(style="thin", color="BFBFBF")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    # Student section header
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=missing_summary.shape[1])
    ws.cell(1, 1).value = "Missing Student Information"

    # Staff section header: Site sits outside the grouped title, like the example screenshot.
    staff_start_col = 9  # I
    ws.cell(1, staff_start_col).value = ""
    ws.merge_cells(start_row=1, start_column=staff_start_col + 1, end_row=1, end_column=staff_start_col + staff_missing_summary.shape[1] - 1)
    ws.cell(1, staff_start_col + 1).value = "Missing Staff Information"

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row in [1, 2]:
                cell.font = Font(name="Arial Narrow", size=10, bold=True)
                cell.fill = blue_fill

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 24 if col in [1, 9] else 17
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 40


def apply_highlights_to_workbook(wb, total_missing_rows, flag_cols, young_dob_rows,
                                 staff_missing_rows=None, staff_flag_cols=None, site_tables=None):
    red_fill = PatternFill("solid", start_color="FF9999", end_color="FF9999")
    blue_fill = PatternFill("solid", start_color="9999FF", end_color="9999FF")

    flag_to_original = {
        fc: fc[: -len("_missing")]
        for fc in flag_cols
        if fc.endswith("_missing") and fc != "State ParticipantID_missing"
    }

    # Student missing pull-out sheets.
    for site, group in total_missing_rows.groupby("Site"):
        sheet_name = clean_sheet_name("Missing - " + str(site), fallback="Missing Site")
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = {cell.value: cell.column for cell in ws[1]}
        for row_idx, (_, row) in enumerate(group.iterrows(), start=2):
            for flag_col, orig_col in flag_to_original.items():
                if orig_col in header and row.get(flag_col, 0) == 1:
                    ws.cell(row=row_idx, column=header[orig_col]).fill = red_fill

    # Staff missing pull-out sheet.
    if staff_missing_rows is not None and staff_flag_cols is not None and "Pull out - Missing Staff Info" in wb.sheetnames:
        ws_staff = wb["Pull out - Missing Staff Info"]
        staff_header = {cell.value: cell.column for cell in ws_staff[1]}
        staff_flag_to_original = {
            "Staff Type_missing": "Staff Type",
            "Employment Type_missing": "Compensation Type",
            "Funded by 21st CCLC_missing": "Funder",
        }
        for row_idx, (_, row) in enumerate(staff_missing_rows.iterrows(), start=2):
            for flag_col in staff_flag_cols:
                orig_col = staff_flag_to_original.get(flag_col, flag_col.replace("_missing", ""))
                if orig_col in staff_header and row.get(flag_col, 0) == 1:
                    ws_staff.cell(row=row_idx, column=staff_header[orig_col]).fill = red_fill

    # Young DOB sheet.
    if "Pull out - Young DOB" in wb.sheetnames:
        ws2 = wb["Pull out - Young DOB"]
        header2 = {cell.value: cell.column for cell in ws2[1]}
        if "Date Of Birth" in header2:
            dob_col_idx = header2["Date Of Birth"]
            for row_idx in range(2, len(young_dob_rows) + 2):
                ws2.cell(row=row_idx, column=dob_col_idx).fill = blue_fill

    # Missing cells in each site summary report.
    if site_tables is not None:
        for site_name in site_tables:
            safe_sheet_name = clean_sheet_name(site_name, fallback="Site")
            if safe_sheet_name not in wb.sheetnames:
                continue
            ws_site = wb[safe_sheet_name]
            for row_idx in range(2, ws_site.max_row + 1):
                for col_idx in range(1, ws_site.max_column + 1):
                    cell = ws_site.cell(row=row_idx, column=col_idx)
                    if cell.value is None or str(cell.value).strip() in ["", "-"]:
                        cell.fill = red_fill


def copy_sheet(src_ws, dest_wb, dest_name, skip_rows=0):
    dest_name = clean_sheet_name(dest_name)
    if dest_name in dest_wb.sheetnames:
        del dest_wb[dest_name]
    dest_ws = dest_wb.create_sheet(title=dest_name)
    for row in src_ws.iter_rows(min_row=skip_rows + 1):
        for cell in row:
            dest_row = cell.row - skip_rows
            dest_cell = dest_ws.cell(row=dest_row, column=cell.column, value=cell.value)
            if cell.has_style:
                dest_cell.font = copy(cell.font)
                dest_cell.border = copy(cell.border)
                dest_cell.fill = copy(cell.fill)
                dest_cell.number_format = cell.number_format
                dest_cell.alignment = copy(cell.alignment)

    for key, dim in src_ws.column_dimensions.items():
        dest_ws.column_dimensions[key].width = dim.width
        dest_ws.column_dimensions[key].hidden = dim.hidden

    for key, dim in src_ws.row_dimensions.items():
        if dim.index > skip_rows:
            dest_ws.row_dimensions[dim.index - skip_rows].height = dim.height
            dest_ws.row_dimensions[dim.index - skip_rows].hidden = dim.hidden


def autofit_basic(wb):
    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 0
            for row in range(1, min(ws.max_row, 80) + 1):
                val = ws.cell(row, col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 35)


# ─────────────────────────────────────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Weekly Update Report Tool", page_icon="📊", layout="centered")
st.title("📊 Weekly Update Report Tool")
st.markdown("Upload your three Excel files, set your target numbers, and download the report — no coding needed.")
st.divider()

st.subheader("Step 1 · Upload Your Excel Files")
col1, col2, col3 = st.columns(3)
with col1:
    students_file = st.file_uploader("🎒 Students File", type=["xlsx", "xls"], key="students")
with col2:
    adults_file = st.file_uploader("👪 Adults File", type=["xlsx", "xls"], key="adults")
with col3:
    all_file = st.file_uploader("📋 Grant Level / All Reports File", type=["xlsx", "xls"], key="all")

st.divider()
institution_name = st.text_input("🏫 Institution Name (used in output filename)", value="Institution")

st.divider()
st.subheader("Step 2 · Set Target Enrollment per Site")
st.info("Enter the target number of students for each site — in the same order as they appear in your Students file. Do not include the Total row.")

if "num_sites" not in st.session_state:
    st.session_state.num_sites = 3

col_add, col_remove = st.columns([1, 1])
with col_add:
    if st.button("➕ Add a site"):
        st.session_state.num_sites += 1
with col_remove:
    if st.button("➖ Remove last site") and st.session_state.num_sites > 1:
        st.session_state.num_sites -= 1

defaults = [152, 200, 100]
target_values = []
for i in range(st.session_state.num_sites):
    default = defaults[i] if i < len(defaults) else 100
    val = st.number_input(f"Site {i + 1} — Target # of students", min_value=1, value=default, key=f"target_{i}")
    target_values.append(val)

st.divider()
st.subheader("Step 3 · Generate Report")
all_uploaded = students_file and adults_file and all_file
if not all_uploaded:
    st.warning("Please upload all three files above before generating.")

if st.button("🚀 Generate Report", disabled=not all_uploaded, type="primary", use_container_width=True):
    with st.status("Running pipeline…", expanded=True) as status:
        try:
            st.write("📂 Reading uploaded files…")
            students_bytes = students_file.read()
            adults_bytes = adults_file.read()
            all_bytes = all_file.read()

            students_xl = pd.ExcelFile(io.BytesIO(students_bytes))
            all_xl = pd.ExcelFile(io.BytesIO(all_bytes))

            # Student Summary Statistics
            st.write("🍀 Processing Student Summary Statistics…")
            part_sheet = find_sheet_name(students_xl, ["participants", "hour"], fallback_index=1)
            df_part_by_hour = clean_participants_by_hour(students_xl.parse(part_sheet))

            daily_sheet = find_sheet_name(students_xl, ["daily", "site", "attendance"], fallback_index=0)
            daily_site_att = students_xl.parse(daily_sheet)
            daily_site_att = set_header_from_row(daily_site_att, 2)
            daily_site_att = daily_site_att[["Total"]].iloc[:-1].copy()
            daily_site_att["Total"] = daily_site_att["Total"].astype(str).str.extract(r"(\d+\.?\d*)")

            # Handle both final notebook column names and older app column names.
            column_aliases = {
                "0 Hours": ["0 Hours", "0"],
                "Less Than 15 Hours": ["Less Than 15 Hours", "Less Than 15"],
                "15-44 Hours": ["15-44 Hours", "15-44"],
                "45-89 Hours": ["45-89 Hours", "45-89"],
                "90-179 Hours": ["90-179 Hours", "90-179"],
                "180-269 Hours": ["180-269 Hours", "180-269"],
                "270+ Hours": ["270+ Hours", "270+"],
            }
            rename_map = {}
            for canonical, aliases in column_aliases.items():
                for alias in aliases:
                    if alias in df_part_by_hour.columns:
                        rename_map[alias] = canonical
                        break
            df_part_by_hour = df_part_by_hour.rename(columns=rename_map)

            all_cols = ["0 Hours", "Less Than 15 Hours", "15-44 Hours", "45-89 Hours", "90-179 Hours", "180-269 Hours", "270+ Hours"]
            served_cols = ["Less Than 15 Hours", "15-44 Hours", "45-89 Hours", "90-179 Hours", "180-269 Hours", "270+ Hours"]
            plus15_cols = ["15-44 Hours", "45-89 Hours", "90-179 Hours", "180-269 Hours", "270+ Hours"]
            plus90_cols = ["90-179 Hours", "180-269 Hours", "270+ Hours"]
            existing_cols = df_part_by_hour.columns.tolist()
            all_cols = [c for c in all_cols if c in existing_cols]
            served_cols = [c for c in served_cols if c in existing_cols]
            plus15_cols = [c for c in plus15_cols if c in existing_cols]
            plus90_cols = [c for c in plus90_cols if c in existing_cols]

            df_part_by_hour[all_cols] = df_part_by_hour[all_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
            df_calc_rows = df_part_by_hour[df_part_by_hour["Site"].astype(str).str.strip().str.lower().ne("total")].copy()
            if "Subtotal" in df_calc_rows["Site"].astype(str).values:
                df_calc_rows = df_calc_rows[df_calc_rows["Site"].astype(str).str.strip() != "Subtotal"]

            df_totals = pd.DataFrame({
                "[Total # Enrolled]": df_calc_rows[all_cols].sum(axis=1),
                "[Total # Served]": df_calc_rows[served_cols].sum(axis=1),
                "[Total # 15+]": df_calc_rows[plus15_cols].sum(axis=1),
                "[Total # 90+]": df_calc_rows[plus90_cols].sum(axis=1),
            }).reset_index(drop=True)

            if len(target_values) != len(df_totals):
                raise ValueError(f"You entered {len(target_values)} target values, but the Students file has {len(df_totals)} site rows. Please adjust the number of sites.")
            df_totals.insert(0, "[Target # of students served]", target_values)

            daily_site_att["Total"] = pd.to_numeric(daily_site_att["Total"], errors="coerce").fillna(0).astype(int)
            if len(daily_site_att) < len(df_totals):
                raise ValueError("Daily Site Attendance Summary has fewer site rows than the student summary.")
            df_totals.insert(3, "Avg. # of Students Per Day", daily_site_att["Total"].iloc[:len(df_totals)].values)

            school_names = [s for s in df_calc_rows["Site"].tolist() if pd.notna(s) and str(s).strip().lower() not in ["", "nan", "subtotal", "total"]]
            df_totals.insert(0, "School", school_names[:len(df_totals)])

            total_row = pd.DataFrame(df_totals.iloc[:, 1:].sum(numeric_only=True)).T
            total_row.insert(0, "School", "Total")
            df_totals = pd.concat([df_totals, total_row], ignore_index=True)

            df_totals["# of students 15+ hrs total (% of Target)"] = (
                df_totals["[Total # 15+]"].astype(int).astype(str)
                + " ("
                + ((df_totals["[Total # 15+]"] / df_totals["[Target # of students served]"]) * 100).round().fillna(0).astype(int).astype(str)
                + "%)"
            )
            df_totals["# of students 90+ hrs total (% of Target)"] = (
                df_totals["[Total # 90+]"].astype(int).astype(str)
                + " ("
                + ((df_totals["[Total # 90+]"] / df_totals["[Target # of students served]"]) * 100).round().fillna(0).astype(int).astype(str)
                + "%)"
            )
            df_totals = df_totals.drop(columns=["[Total # 15+]", "[Total # 90+]"]).reset_index(drop=True)
            styled_totals = df_totals.style.applymap(color_pct_cols, subset=["# of students 15+ hrs total (% of Target)"])

            # Family Component
            st.write("🌷 Processing Family Component…")
            df_hours = pd.read_excel(io.BytesIO(adults_bytes), sheet_name=find_sheet_name(pd.ExcelFile(io.BytesIO(adults_bytes)), ["participant", "attendance", "hours"], fallback_index=2), skiprows=2)
            df_hours["HoursPresent"] = pd.to_numeric(df_hours["HoursPresent"], errors="coerce")
            df_hours["ParticipantId"] = df_hours["ParticipantId"].astype(str).str.replace(r"\.0$", "", regex=True)
            df_active = df_hours[(df_hours["HoursPresent"] > 0) & (df_hours["ParticipantId"].str.len() != 9)]
            result = df_active.groupby("Site")["ParticipantId"].nunique().reset_index()
            result.rename(columns={"ParticipantId": "Parents Served (Total)"}, inplace=True)
            result.loc[len(result)] = {"Site": "Total", "Parents Served (Total)": result["Parents Served (Total)"].sum()}

            # Student missing information
            st.write("🌸 Processing Participant Demographics…")
            demo_sheet = find_sheet_name(students_xl, ["participant", "demographics"], fallback_index=4)
            df_part_demo = set_header_from_row(students_xl.parse(demo_sheet), 2)
            missing_summary, missing_site_rows, total_missing_rows, flag_cols, young_dob_rows = summarize_missing_by_school(
                df_part_demo, ["Date Of Birth", "Grade Level", "Gender"]
            )

            # Staff missing information
            st.write("🧑‍🏫 Processing Staff Details…")
            staff_sheet = find_sheet_name(all_xl, ["staff", "details"])
            df_staff = pd.read_excel(io.BytesIO(all_bytes), sheet_name=staff_sheet, skiprows=2)
            staff_missing_summary, staff_missing_rows, staff_flag_cols = summarize_staff_missing_info(df_staff)

            # Site Summary Reports
            st.write("🪻 Building Site Summary Reports…")
            act_sheet = find_sheet_name(all_xl, ["activity", "session", "details"], fallback_index=3)
            enr_sheet = find_sheet_name(all_xl, ["session", "enrollment"], fallback_index=5)
            att_sheet = find_sheet_name(all_xl, ["daily", "activity", "attendance"], fallback_index=6)

            df_act = pd.read_excel(io.BytesIO(all_bytes), sheet_name=act_sheet, skiprows=2)
            df_enr = pd.read_excel(io.BytesIO(all_bytes), sheet_name=enr_sheet, skiprows=2)
            df_att = pd.read_excel(io.BytesIO(all_bytes), sheet_name=att_sheet, skiprows=4)
            if "Site" not in df_att.columns:
                df_att_raw = pd.read_excel(io.BytesIO(all_bytes), sheet_name=att_sheet)
                df_att = set_header_from_row(df_att_raw, 3)

            df_act = df_act[["Site", "Activity", "Session", "Days Scheduled", "Session Start Date"]].copy()
            df_act["Days Scheduled"] = pd.to_numeric(df_act["Days Scheduled"], errors="coerce")
            df_act["Session Start Date"] = pd.to_datetime(df_act["Session Start Date"], errors="coerce")
            today = pd.Timestamp.today().normalize()

            df_enr = df_enr[["Site", "Activity", "Session", "Enrolled Count"]].copy()
            df_enr["Enrolled Count"] = pd.to_numeric(df_enr["Enrolled Count"], errors="coerce")
            df_enr.rename(columns={"Enrolled Count": "Enrolled Participant"}, inplace=True)

            df_att = df_att[["Site", "Activity", "Session", "Total"]].copy()
            df_att["Total"] = df_att["Total"].apply(lambda val: np.nan if pd.isna(val) else pd.to_numeric(str(val).replace("Average:", "").strip(), errors="coerce")).round(0)
            df_att.rename(columns={"Total": "Average Daily Attendance"}, inplace=True)

            sites = [
                s for s in df_act["Site"].dropna().unique().tolist()
                if str(s).strip() != "" and not str(s).startswith("Total") and not str(s).startswith("Average")
            ]
            site_tables = {}
            for site in sites:
                merged_site = pd.merge(
                    df_act[df_act["Site"] == site],
                    df_enr[df_enr["Site"] == site],
                    on=["Site", "Activity", "Session"],
                    how="outer",
                )
                merged_site = pd.merge(
                    merged_site,
                    df_att[df_att["Site"] == site],
                    on=["Site", "Activity", "Session"],
                    how="outer",
                )
                merged_site = merged_site[~(merged_site["Session Start Date"] >= today)]
                merged_site = merged_site.drop(columns=["Session Start Date"], errors="ignore")
                merged_site = merged_site.fillna("-")
                site_tables[site] = merged_site.sort_values(by=["Session"], ascending=True).reset_index(drop=True)

            # Write Excel
            st.write("🌈 Writing Excel report…")
            output_buffer = io.BytesIO()
            hide_cols = {"Race/Ethnicity", "English Learner Status", "Lunch Status", "Special Education Status", "IDEA Disability Type"}

            with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
                styled_totals.to_excel(writer, sheet_name="Student Summary Statistics", index=False)
                result.to_excel(writer, sheet_name="Family Component Summary", index=False)
                write_combined_missing_summary(writer, missing_summary, staff_missing_summary)
                missing_summary.to_excel(writer, sheet_name="Missing Student Summary", index=False)
                staff_missing_summary.to_excel(writer, sheet_name="Missing Staff Summary", index=False)
                missing_site_rows.to_excel(writer, sheet_name="Pull out - Missing Site Info", index=False)

                display_cols = [c for c in total_missing_rows.columns if not c.endswith("_missing") and c != "DOB_too_young" and c not in hide_cols]
                for site, group in total_missing_rows.groupby("Site"):
                    safe_name = clean_sheet_name("Missing - " + str(site), fallback="Missing Site")
                    group[display_cols].to_excel(writer, sheet_name=safe_name, index=False)

                young_display_cols = [c for c in young_dob_rows.columns if not c.endswith("_missing") and c != "DOB_too_young"]
                young_dob_rows[young_display_cols].to_excel(writer, sheet_name="Pull out - Young DOB", index=False)

                staff_display_cols = [c for c in staff_missing_rows.columns if not c.endswith("_missing") and c != "_staff_id"]
                staff_missing_rows[staff_display_cols].to_excel(writer, sheet_name="Pull out - Missing Staff Info", index=False)

                for site_name, final_df in site_tables.items():
                    safe_sheet_name = clean_sheet_name(site_name, fallback="Site")
                    final_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

            output_buffer.seek(0)
            wb = load_workbook(output_buffer)
            apply_highlights_to_workbook(
                wb,
                total_missing_rows,
                flag_cols,
                young_dob_rows,
                staff_missing_rows=staff_missing_rows,
                staff_flag_cols=staff_flag_cols,
                site_tables=site_tables,
            )

            # Copy source sheets
            st.write("📋 Copying source sheets…")
            students_wb_src = load_workbook(io.BytesIO(students_bytes))
            adults_wb_src = load_workbook(io.BytesIO(adults_bytes))
            copy_sheet(students_wb_src[part_sheet], wb, "Students - Participants By Hour", skip_rows=0)
            copy_sheet(adults_wb_src[find_sheet_name(adults_wb_src, ["participants", "hour"], fallback_index=1)], wb, "Adults - Participants By Hour", skip_rows=0)
            copy_sheet(students_wb_src[demo_sheet], wb, "Participant Demographics", skip_rows=0)

            autofit_basic(wb)
            arial_narrow = Font(name="Arial Narrow", size=10)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        # Keep bold/coloring where already applied by preserving bold and fill.
                        existing_bold = cell.font.bold
                        cell.font = Font(name="Arial Narrow", size=10, bold=existing_bold)

            final_buffer = io.BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)

            status.update(label="✅ Report ready!", state="complete")
            st.session_state.output_bytes = final_buffer.read()
            st.session_state.report_ready = True

        except Exception as e:
            status.update(label="❌ Something went wrong", state="error")
            st.error(f"Error: {e}")
            st.session_state.report_ready = False

if st.session_state.get("report_ready"):
    safe_institution = institution_name.strip().replace(" ", "_") or "Institution"
    today_str = date.today().strftime("%Y%m%d")
    output_filename = f"{today_str}_{safe_institution}_WeeklyUpdates.xlsx"
    st.success("🎉 Your report is ready!")
    st.download_button(
        label=f"⬇️ Download {output_filename}",
        data=st.session_state.output_bytes,
        file_name=output_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )
